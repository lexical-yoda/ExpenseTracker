import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import json
import threading
from datetime import date, datetime
from collections import defaultdict

# File lock for xlsx read-modify-write operations
_xlsx_lock = threading.Lock()
_accounts_lock = threading.Lock()

# ── Config ─────────────────────────────────────────────────────────────────────

DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')
XLSX_PATH = os.environ.get('EXPENSES_XLSX', os.path.join(DATA_DIR, 'expenses.xlsx'))
ACCOUNTS_FILE = os.path.join(DATA_DIR, 'accounts.json')

COLUMNS = {
    'date': 1,          # A
    'txn_id': 2,        # B
    'description': 3,   # C
    'category': 4,      # D
    'sub_category': 5,  # E
    'account': 6,       # F — stores full account name
    'amount': 7,        # G
    'parent_id': 8,     # H
    'txn_type': 9,      # I
    'track': 10,        # J — Yes/No, controls dashboard visibility
    'units': 11,        # K — units bought/sold for investment accounts
    'emi_id': 12,       # L — links installment transactions to an EMI account
}

TABLE_START = 1  # Column headers row
DATA_START = 2   # First transaction row


# ── Read cache ─────────────────────────────────────────────────────────────────
# Cache parsed transactions to avoid re-reading xlsx on every request.
# Invalidated when file modification time changes (i.e., after any write).
_txn_cache = {'mtime': 0, 'data': None}
_cache_lock = threading.Lock()


# ── Workbook helpers ───────────────────────────────────────────────────────────

def _invalidate_cache():
    with _cache_lock:
        _txn_cache['mtime'] = 0
        _txn_cache['data'] = None

def load_workbook(data_only=False):
    if not os.path.exists(XLSX_PATH):
        wb = openpyxl.Workbook()
        wb.active.title = '_init'
        wb.save(XLSX_PATH)
    return openpyxl.load_workbook(XLSX_PATH, data_only=data_only)


def save_workbook(wb):
    wb.save(XLSX_PATH)
    _invalidate_cache()


def month_sheet_name(year=None, month=None):
    if year is None or month is None:
        today = date.today()
        year, month = today.year, today.month
    return datetime(year, month, 1).strftime('%B %Y')


def ensure_month_sheet(year=None, month=None):
    """Create month sheet if it doesn't exist. Returns (wb, ws)."""
    wb = load_workbook()
    name = month_sheet_name(year, month)

    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        _init_sheet(ws)
        if '_init' in wb.sheetnames:
            del wb['_init']
        save_workbook(wb)
        wb = load_workbook()

    return wb, wb[name]


def _init_sheet(ws):
    """Set up headers and formatting for a new month sheet."""
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['Date', 'Txn ID', 'Description', 'Category', 'Sub-Category',
               'Account', 'Amount (₹)', 'Parent ID', 'Type', 'Track', 'Units', 'EMI ID']
    widths = [12, 8, 28, 16, 16, 24, 14, 10, 10, 8, 10, 8]

    for col, header in enumerate(headers, start=1):
        if not header:
            continue
        cell = ws.cell(row=TABLE_START, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='1a1a2e')
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Transaction ID management ─────────────────────────────────────────────────

def get_next_txn_id(wb):
    """Scan all sheets and return max txn_id + 1."""
    max_id = 0
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=DATA_START, values_only=True):
            txn_id = row[COLUMNS['txn_id'] - 1]
            if isinstance(txn_id, int) and txn_id > max_id:
                max_id = txn_id
    return max_id + 1


def sanitize_cell(value):
    """Prevent formula injection in spreadsheet cells."""
    if isinstance(value, str) and value and value[0] in ('=', '+', '-', '@', '|', '\t'):
        return "'" + value
    return value


# ── Read ───────────────────────────────────────────────────────────────────────

def parse_row(row, sheet_name):
    """Convert a raw openpyxl row tuple to a dict."""
    def val(col_name):
        idx = COLUMNS[col_name] - 1
        return row[idx] if idx < len(row) else None

    txn_id = val('txn_id')
    if not isinstance(txn_id, int):
        return None

    date_val = val('date')
    if isinstance(date_val, datetime):
        date_str = date_val.strftime('%Y-%m-%d')
    elif isinstance(date_val, date):
        date_str = date_val.strftime('%Y-%m-%d')
    else:
        date_str = str(date_val) if date_val else ''

    txn_type = val('txn_type')
    track_val = val('track')
    # Default to 'Yes' if column is empty or missing (backward compat)
    tracked = True if track_val is None or str(track_val).strip().lower() in ('yes', 'true', '1', '') else False

    units_val = val('units')
    units = float(units_val) if units_val is not None and units_val != '' else None

    emi_id_val = val('emi_id')
    emi_id = int(emi_id_val) if isinstance(emi_id_val, int) or (isinstance(emi_id_val, str) and emi_id_val.isdigit()) else None

    return {
        'id': txn_id,
        'date': date_str,
        'description': val('description') or '',
        'category': val('category') or '',
        'sub_category': val('sub_category') or '',
        'account': val('account') or '',
        'amount': float(val('amount') or 0),
        'parent_id': val('parent_id'),
        'type': txn_type or 'Expense',
        'track': tracked,
        'units': units,
        'emi_id': emi_id,
        'sheet': sheet_name,
    }


def get_all_transactions():
    """Return all transactions across all sheets, sorted by date desc. Cached."""
    with _cache_lock:
        if os.path.exists(XLSX_PATH):
            mtime = os.path.getmtime(XLSX_PATH)
            if _txn_cache['data'] is not None and _txn_cache['mtime'] == mtime:
                return _txn_cache['data']

    wb = load_workbook(data_only=True)
    transactions = []
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=DATA_START, values_only=True):
            parsed = parse_row(row, name)
            if parsed:
                transactions.append(parsed)
    transactions.sort(key=lambda t: (t['date'], t['id']), reverse=True)

    with _cache_lock:
        if os.path.exists(XLSX_PATH):
            _txn_cache['mtime'] = os.path.getmtime(XLSX_PATH)
            _txn_cache['data'] = transactions

    return transactions


def get_transaction_by_id(txn_id):
    wb = load_workbook(data_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=DATA_START, values_only=True):
            parsed = parse_row(row, name)
            if parsed and parsed['id'] == txn_id:
                return parsed
    return None


# ── Write ──────────────────────────────────────────────────────────────────────

def add_transaction(date_str, description, category, sub_category, account, amount, parent_id=None, txn_type='Expense', track=True, units=None, emi_id=None):
    """Append a transaction to the correct month sheet. Returns txn_id."""
    with _xlsx_lock:
        d = datetime.strptime(date_str, '%Y-%m-%d').date()
        wb, ws = ensure_month_sheet(d.year, d.month)

        txn_id = get_next_txn_id(wb)
        sheet_name = month_sheet_name(d.year, d.month)
        ws_fresh = wb[sheet_name]

        # Find next empty row
        next_row = DATA_START
        while ws_fresh.cell(row=next_row, column=COLUMNS['txn_id']).value is not None:
            next_row += 1

        # Write row
        ws_fresh.cell(row=next_row, column=COLUMNS['date']).value = d
        ws_fresh.cell(row=next_row, column=COLUMNS['txn_id']).value = txn_id
        ws_fresh.cell(row=next_row, column=COLUMNS['description']).value = sanitize_cell(description)
        ws_fresh.cell(row=next_row, column=COLUMNS['category']).value = sanitize_cell(category)
        ws_fresh.cell(row=next_row, column=COLUMNS['sub_category']).value = sanitize_cell(sub_category or '')
        ws_fresh.cell(row=next_row, column=COLUMNS['account']).value = sanitize_cell(account)
        ws_fresh.cell(row=next_row, column=COLUMNS['amount']).value = amount
        ws_fresh.cell(row=next_row, column=COLUMNS['parent_id']).value = parent_id
        ws_fresh.cell(row=next_row, column=COLUMNS['txn_type']).value = txn_type
        ws_fresh.cell(row=next_row, column=COLUMNS['track']).value = 'Yes' if track else 'No'
        if units is not None:
            ws_fresh.cell(row=next_row, column=COLUMNS['units']).value = units
        if emi_id is not None:
            ws_fresh.cell(row=next_row, column=COLUMNS['emi_id']).value = int(emi_id)
        ws_fresh.cell(row=next_row, column=COLUMNS['amount']).number_format = '₹#,##0.00'

        save_workbook(wb)

        # Auto-update investment account units & balance (inside lock for consistency)
        if units is not None and not parent_id:
            _update_investment_account(account, amount, units, txn_type)

    return txn_id


def _update_investment_account(account_name, amount, units, txn_type):
    """Auto-update investment account units and invested amount on transaction."""
    with _accounts_lock:
        if not os.path.exists(ACCOUNTS_FILE):
            return
        with open(ACCOUNTS_FILE, 'r') as f:
            accounts = json.load(f)

        for acct in accounts:
            if acct['name'] == account_name and acct.get('type') == 'investment':
                if txn_type == 'Income':
                    acct['units'] = acct.get('units', 0) + units
                    acct['balance'] = acct.get('balance', 0) + amount
                elif txn_type == 'Expense':
                    acct['units'] = max(0, acct.get('units', 0) - units)
                    acct['balance'] = max(0, acct.get('balance', 0) - amount)
                import tempfile
                fd, tmp_path = tempfile.mkstemp(dir=os.path.dirname(ACCOUNTS_FILE), suffix='.tmp')
                try:
                    with os.fdopen(fd, 'w') as f:
                        json.dump(accounts, f, indent=2)
                    os.replace(tmp_path, ACCOUNTS_FILE)
                except Exception:
                    try:
                        os.unlink(tmp_path)
                    except OSError:
                        pass
                    raise
                break


# ── Find / Edit / Delete ──────────────────────────────────────────────────────

def find_transaction_row(wb, txn_id):
    """Find the sheet name and row number for a given txn_id."""
    for name in wb.sheetnames:
        ws = wb[name]
        for row_num in range(DATA_START, ws.max_row + 1):
            if ws.cell(row=row_num, column=COLUMNS['txn_id']).value == txn_id:
                return (name, row_num)
    return None


def update_transaction(txn_id, data):
    """Update a transaction in place. Returns updated dict."""
    with _xlsx_lock:
        return _update_transaction_inner(txn_id, data)

def _update_transaction_inner(txn_id, data):
    wb = load_workbook()
    result = find_transaction_row(wb, txn_id)
    if not result:
        raise ValueError(f'Transaction {txn_id} not found')

    sheet_name, row_num = result
    ws = wb[sheet_name]

    old_date = ws.cell(row=row_num, column=COLUMNS['date']).value
    new_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
    if isinstance(old_date, datetime):
        old_date = old_date.date()
    elif isinstance(old_date, str):
        old_date = datetime.strptime(old_date, '%Y-%m-%d').date()
    elif old_date is None:
        raise ValueError(f'Transaction {txn_id} has no date')

    if (old_date.year, old_date.month) != (new_date.year, new_date.month):
        # Cross-month edit: delete from old sheet and add to new sheet
        parent_id = ws.cell(row=row_num, column=COLUMNS['parent_id']).value
        ws.delete_rows(row_num)
        save_workbook(wb)
        _invalidate_cache()

        # Re-add to the correct month sheet
        wb2, ws2 = ensure_month_sheet(new_date.year, new_date.month)
        next_row = DATA_START
        while ws2.cell(row=next_row, column=COLUMNS['txn_id']).value is not None:
            next_row += 1
        ws2.cell(row=next_row, column=COLUMNS['date']).value = new_date
        ws2.cell(row=next_row, column=COLUMNS['txn_id']).value = txn_id
        ws2.cell(row=next_row, column=COLUMNS['description']).value = sanitize_cell(data['description'])
        ws2.cell(row=next_row, column=COLUMNS['category']).value = sanitize_cell(data['category'])
        ws2.cell(row=next_row, column=COLUMNS['sub_category']).value = sanitize_cell(data.get('sub_category', ''))
        ws2.cell(row=next_row, column=COLUMNS['account']).value = sanitize_cell(data['account'])
        ws2.cell(row=next_row, column=COLUMNS['amount']).value = float(data['amount'])
        ws2.cell(row=next_row, column=COLUMNS['amount']).number_format = '₹#,##0.00'
        ws2.cell(row=next_row, column=COLUMNS['txn_type']).value = data.get('type', 'Expense')
        if 'track' in data:
            ws2.cell(row=next_row, column=COLUMNS['track']).value = 'Yes' if data['track'] else 'No'
        if 'units' in data and data['units'] is not None:
            ws2.cell(row=next_row, column=COLUMNS['units']).value = float(data['units'])
        if 'emi_id' in data:
            ws2.cell(row=next_row, column=COLUMNS['emi_id']).value = int(data['emi_id']) if data['emi_id'] else None
        if parent_id is not None:
            ws2.cell(row=next_row, column=COLUMNS['parent_id']).value = parent_id
        save_workbook(wb2)
        _invalidate_cache()
        return get_transaction_by_id(txn_id)

    ws.cell(row=row_num, column=COLUMNS['date']).value = new_date
    ws.cell(row=row_num, column=COLUMNS['description']).value = sanitize_cell(data['description'])
    ws.cell(row=row_num, column=COLUMNS['category']).value = sanitize_cell(data['category'])
    ws.cell(row=row_num, column=COLUMNS['sub_category']).value = sanitize_cell(data.get('sub_category', ''))
    ws.cell(row=row_num, column=COLUMNS['account']).value = sanitize_cell(data['account'])
    ws.cell(row=row_num, column=COLUMNS['amount']).value = float(data['amount'])
    ws.cell(row=row_num, column=COLUMNS['amount']).number_format = '₹#,##0.00'
    ws.cell(row=row_num, column=COLUMNS['txn_type']).value = data.get('type', 'Expense')
    if 'track' in data:
        ws.cell(row=row_num, column=COLUMNS['track']).value = 'Yes' if data['track'] else 'No'
    if 'units' in data and data['units'] is not None:
        ws.cell(row=row_num, column=COLUMNS['units']).value = float(data['units'])
    if 'emi_id' in data:
        ws.cell(row=row_num, column=COLUMNS['emi_id']).value = int(data['emi_id']) if data['emi_id'] else None

    save_workbook(wb)
    return get_transaction_by_id(txn_id)


def delete_transaction(txn_id):
    """Delete a transaction. If parent, also deletes all sub-items in all sheets."""
    with _xlsx_lock:
        return _delete_transaction_inner(txn_id)

def _delete_transaction_inner(txn_id):
    wb = load_workbook()
    result = find_transaction_row(wb, txn_id)
    if not result:
        raise ValueError(f'Transaction {txn_id} not found')

    sheet_name, row_num = result
    ws = wb[sheet_name]

    # Capture investment data before deletion for reversal
    parent_id = ws.cell(row=row_num, column=COLUMNS['parent_id']).value
    units_val = ws.cell(row=row_num, column=COLUMNS['units']).value if COLUMNS['units'] - 1 < ws.max_column else None
    txn_amount = ws.cell(row=row_num, column=COLUMNS['amount']).value
    txn_type = ws.cell(row=row_num, column=COLUMNS['txn_type']).value or 'Expense'
    txn_account = ws.cell(row=row_num, column=COLUMNS['account']).value

    rows_to_delete = [row_num]
    deleted_ids = [txn_id]

    if parent_id is None:
        # Search ALL sheets for children (sub-items may be in different months)
        children_to_delete = {}  # {sheet_name: [row_nums]}
        for sname in wb.sheetnames:
            s = wb[sname]
            for r in range(DATA_START, s.max_row + 1):
                pid = s.cell(row=r, column=COLUMNS['parent_id']).value
                if pid == txn_id:
                    if sname not in children_to_delete:
                        children_to_delete[sname] = []
                    children_to_delete[sname].append(r)
                    deleted_ids.append(s.cell(row=r, column=COLUMNS['txn_id']).value)
        # Delete children from other sheets
        for sname, rows in children_to_delete.items():
            s = wb[sname]
            for r in sorted(rows, reverse=True):
                if sname == sheet_name:
                    rows_to_delete.append(r)
                else:
                    s.delete_rows(r, 1)

    for r in sorted(set(rows_to_delete), reverse=True):
        ws.delete_rows(r, 1)

    save_workbook(wb)

    # Reverse investment account update if this was a parent with units
    if parent_id is None and units_val is not None and txn_account:
        try:
            units = float(units_val)
            amount = float(txn_amount or 0)
            # Reverse: if original was Income (added units), now subtract; vice versa
            reverse_type = 'Expense' if txn_type == 'Income' else 'Income'
            _update_investment_account(txn_account, amount, units, reverse_type)
        except (ValueError, TypeError):
            pass

    return deleted_ids


# ── Account operations ────────────────────────────────────────────────────────

def rename_account_in_sheets(old_name, new_name):
    """Rename an account across all transaction sheets."""
    with _xlsx_lock:
        wb = load_workbook()
        changed = False
        for name in wb.sheetnames:
            ws = wb[name]
            for row_num in range(DATA_START, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=COLUMNS['account'])
                if cell.value == old_name:
                    cell.value = new_name
                    changed = True
        if changed:
            save_workbook(wb)


def compute_emi_schedule(principal, annual_rate, tenure_months, first_installment_date):
    """Compute EMI schedule using reducing-balance method.

    Args:
        principal: float — loan amount
        annual_rate: float — annual interest rate in percent (e.g. 16 for 16%)
        tenure_months: int — number of monthly installments
        first_installment_date: 'YYYY-MM-DD' string OR date object — date of first EMI

    Returns:
        list of dicts, each with: index (1-based), date, principal, interest,
        installment, balance_after. Last installment adjusts for rounding so
        balance_after is exactly 0.
    """
    try:
        principal = float(principal or 0)
        tenure_months = int(tenure_months or 0)
        annual_rate = float(annual_rate or 0)
    except (ValueError, TypeError):
        return []
    if principal <= 0 or tenure_months <= 0:
        return []

    if isinstance(first_installment_date, str):
        try:
            start = datetime.strptime(first_installment_date, '%Y-%m-%d').date()
        except ValueError:
            return []
    elif isinstance(first_installment_date, date):
        start = first_installment_date
    else:
        return []

    r = (annual_rate / 100.0) / 12.0
    if r > 0:
        emi = principal * r * ((1 + r) ** tenure_months) / (((1 + r) ** tenure_months) - 1)
    else:
        emi = principal / tenure_months

    from calendar import monthrange

    schedule = []
    balance = principal
    for i in range(tenure_months):
        year = start.year + (start.month - 1 + i) // 12
        month = (start.month - 1 + i) % 12 + 1
        last_day = monthrange(year, month)[1]
        due_day = min(start.day, last_day)
        due_date = date(year, month, due_day)

        interest_part = balance * r
        if i == tenure_months - 1:
            principal_part = balance
            installment = principal_part + interest_part
        else:
            principal_part = emi - interest_part
            installment = emi
        balance = max(0.0, balance - principal_part)

        schedule.append({
            'index': i + 1,
            'date': due_date.strftime('%Y-%m-%d'),
            'principal': round(principal_part, 2),
            'interest': round(interest_part, 2),
            'installment': round(installment, 2),
            'balance_after': round(balance, 2),
        })
    return schedule


def compute_account_balances():
    """Compute current balance for each account from accounts.json + transactions."""
    with _accounts_lock:
        if not os.path.exists(ACCOUNTS_FILE):
            return []
        with open(ACCOUNTS_FILE, 'r') as f:
            accounts = json.load(f)

    transactions = get_all_transactions()
    parents = [t for t in transactions if not t['parent_id']]

    spend_by_account = defaultdict(float)
    income_by_account = defaultdict(float)
    emi_payments_by_id = defaultdict(list)  # emi_id -> [txn, ...]
    for t in parents:
        if t['type'] == 'Income':
            income_by_account[t['account']] += t['amount']
        else:
            # Both Expense and Transfer reduce account balance
            spend_by_account[t['account']] += t['amount']
        if t.get('emi_id'):
            emi_payments_by_id[int(t['emi_id'])].append(t)

    result = []
    for acct in accounts:
        name = acct['name']
        spent = spend_by_account.get(name, 0)
        earned = income_by_account.get(name, 0)
        if acct['type'] == 'savings':
            current = acct['balance'] - spent + earned
            result.append({**acct, 'current_balance': round(current, 2)})
        elif acct['type'] == 'credit':
            accumulated = spent - earned
            remaining = acct['limit'] - accumulated
            result.append({**acct, 'accumulated': round(accumulated, 2), 'remaining': round(remaining, 2)})
        elif acct['type'] == 'investment':
            # Investment accounts: balance = cost basis / principal
            entry = {**acct, 'invested': acct.get('balance', 0)}
            entry['subtype'] = acct.get('subtype', 'market')
            result.append(entry)
        elif acct['type'] == 'emi':
            principal = float(acct.get('principal', 0) or 0)
            tenure = int(acct.get('tenure_months', 0) or 0)
            rate = float(acct.get('interest_rate', 0) or 0)
            first_date = acct.get('first_installment_date') or ''
            schedule = compute_emi_schedule(principal, rate, tenure, first_date) if first_date else []

            # Count paid installments = transactions with matching emi_id, sorted by date
            paid_txns = sorted(emi_payments_by_id.get(acct['id'], []), key=lambda t: t['date'])
            paid_count = min(len(paid_txns), len(schedule))

            principal_paid = sum(s['principal'] for s in schedule[:paid_count])
            interest_paid = sum(s['interest'] for s in schedule[:paid_count])
            interest_remaining = sum(s['interest'] for s in schedule[paid_count:])
            outstanding = max(0.0, principal - principal_paid)
            next_schedule = schedule[paid_count] if paid_count < len(schedule) else None

            entry = {**acct,
                     'outstanding_principal': round(outstanding, 2),
                     'principal_paid': round(principal_paid, 2),
                     'interest_paid': round(interest_paid, 2),
                     'interest_remaining': round(interest_remaining, 2),
                     'installments_paid': paid_count,
                     'installments_total': len(schedule),
                     'installments_remaining': max(0, len(schedule) - paid_count),
                     'next_due_date': next_schedule['date'] if next_schedule else None,
                     'next_installment': next_schedule['installment'] if next_schedule else 0,
                     'status': 'closed' if paid_count >= len(schedule) and len(schedule) > 0 else 'active',
                     }
            result.append(entry)
    return result


# ── Summary for dashboard ──────────────────────────────────────────────────────

def get_monthly_summary():
    """Return structured data for Plotly dashboard."""
    transactions = get_all_transactions()
    parents = [t for t in transactions if not t['parent_id']]

    monthly = defaultdict(lambda: defaultdict(float))
    daily = defaultdict(float)
    by_category = defaultdict(float)
    by_account = defaultdict(float)

    for t in parents:
        if t['type'] == 'Transfer':
            continue  # Transfers excluded from summary charts
        if t['type'] == 'Income':
            continue  # Income excluded from spending charts
        if not t['track']:
            continue  # Untracked transactions excluded from charts
        month = t['date'][:7]
        monthly[month][t['category']] += t['amount']
        daily[t['date']] += t['amount']
        by_category[t['category']] += t['amount']
        by_account[t['account']] += t['amount']

    return {
        'monthly': {k: dict(v) for k, v in sorted(monthly.items())},
        'daily': dict(sorted(daily.items())),
        'by_category': dict(sorted(by_category.items(), key=lambda x: -x[1])),
        'by_account': dict(by_account),
        'transactions': parents,
    }
